import sys, json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

data = json.load(open(sys.argv[1]))
pemasukan   = data['pemasukan']
pengeluaran = data['pengeluaran']
nama_user   = data['nama_user']
tgl_cetak   = data['tanggal_cetak']

total_p  = sum(int(r['jumlah']) for r in pemasukan)
total_px = sum(int(r['jumlah']) for r in pengeluaran)

wb = Workbook()
ws = wb.active
ws.title = "Laporan Keuangan"

BOLD = Font(name="Arial", bold=True, size=11)
NORM = Font(name="Arial", size=11)
ITAL = Font(name="Arial", italic=True, size=10)

def w(row, col, val, font=None, align="left", num_fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font or NORM
    c.alignment = Alignment(horizontal=align)
    if num_fmt:
        c.number_format = num_fmt
    return c

row = 1
w(row, 1, "LAPORAN KEUANGAN HMJ TI NAKAMA", BOLD)
row += 1
w(row, 1, f"Periode: {datetime.now().strftime('%d-%m-%Y')}")
row += 2

# PEMASUKAN
w(row, 1, "DETAIL PEMASUKAN", BOLD); row += 1
for col, label, align in [(1,"NO.","left"),(2,"TANGGAL","left"),(3,"SUMBER DANA","left"),(4,"NAMA PENCATAT","left"),(5,"JUMLAH (Rp)","left"),(6,"STATUS","left")]:
    w(row, col, label, BOLD, align)
row += 1

for i, r in enumerate(pemasukan, 1):
    w(row, 1, i)
    w(row, 2, r['tanggal'])
    w(row, 3, r['sumber_dana'])
    w(row, 4, r['nama_lengkap'])
    w(row, 5, int(r['jumlah']), num_fmt='#,##0')
    w(row, 6, r['status'])
    row += 1

if not pemasukan:
    w(row, 4, "Tidak ada data"); row += 1

w(row, 4, "TOTAL PEMASUKAN", BOLD)
w(row, 5, total_p, BOLD, num_fmt='#,##0')
row += 3

# PENGELUARAN
w(row, 1, "DETAIL PENGELUARAN", BOLD); row += 1
for col, label in [(1,"NO."),(2,"TANGGAL"),(3,"PENERIMA"),(4,"NAMA PENCATAT"),(5,"JUMLAH (Rp)"),(6,"STATUS")]:
    w(row, col, label, BOLD)
row += 1

for i, r in enumerate(pengeluaran, 1):
    w(row, 1, i)
    w(row, 2, r['tanggal'])
    w(row, 3, r['penerima'])
    w(row, 4, r['nama_lengkap'])
    w(row, 5, int(r['jumlah']), num_fmt='#,##0')
    w(row, 6, r['status'])
    row += 1

if not pengeluaran:
    w(row, 4, "Tidak ada data"); row += 1

w(row, 4, "TOTAL PENGELUARAN", BOLD)
w(row, 5, total_px, BOLD, num_fmt='#,##0')
row += 3

# RINGKASAN
w(row, 1, "RINGKASAN KEUANGAN", BOLD); row += 1
w(row, 1, "KETERANGAN", BOLD); w(row, 2, "JUMLAH (Rp)", BOLD); row += 1
w(row, 1, "Total Pemasukan"); w(row, 2, total_p, num_fmt='#,##0'); row += 1
w(row, 1, "Total Pengeluaran"); w(row, 2, total_px, num_fmt='#,##0'); row += 1
w(row, 1, "SALDO AKHIR", BOLD); w(row, 2, total_p - total_px, BOLD, num_fmt='#,##0'); row += 2

w(row, 1, f"Dicetak pada: {tgl_cetak}", ITAL); row += 1
w(row, 1, f"Oleh: {nama_user}", ITAL)

# Auto-fit lebar kolom
col_widths = {1:6, 2:16, 3:30, 4:25, 5:18, 6:14}
for col, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

wb.save(sys.argv[2])
